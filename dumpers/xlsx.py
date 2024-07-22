from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

def update_xlsx(filename, symbol, timestamp, underlying_value, transformed_data):
    """Update or create Excel file with option chain data."""
    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active
    ws.title = 'Options Data'

    # Find the column where new data should start
    column_number = find_column_to_update(ws, timestamp)

    kwargs = {
        'transformed_data': transformed_data,
        'timestamp': timestamp,
        'underlying_value': underlying_value,
        'column_number': column_number
    }

    if column_number is None:
        column_number = find_empty_column(ws)
        initialize_sheet(ws, symbol, **kwargs)
    else:
        add_new_lastprice(ws, **kwargs)

    # Adjust column widths based on content
    adjust_column_widths(ws)

    # Save workbook
    wb.save(filename)
    print(f"XLSX file '{filename}' successfully updated.")

def find_column_to_update(ws, timestamp):
    """Find the column index to update based on timestamp."""
    max_col = ws.max_column
    for col in range(2, max_col + 1):
        cell_value = ws.cell(row=4, column=col).value
        if cell_value == f'DATE {timestamp.strftime("%d-%m-%Y")}':
            return col
    return None

def find_empty_column(ws):
    """Find the first empty column in the worksheet."""
    max_col = ws.max_column
    for col in range(2, max_col + 2):
        if ws.cell(row=3, column=col).value is None:
            return col
    return max_col + 1

def initialize_sheet(ws, symbol, transformed_data, timestamp, underlying_value, column_number):
    """Initialize the worksheet with headers and timestamp."""
    ws.cell(row=3, column=column_number).value = symbol
    ws.cell(row=3, column=column_number).font = Font(bold=True)
    ws.cell(row=4, column=column_number).value = f'DATE {timestamp.strftime("%d-%m-%Y")}'
    ws.cell(row=4, column=column_number).font = Font(bold=True)

    headers = ['Identifier', 'Strike Price / Type', timestamp.strftime('%H:%M')]
    for i, header in enumerate(headers, start=1):
        ws.cell(row=5, column=column_number + i - 1).value = header
        ws.cell(row=5, column=column_number + i - 1).font = Font(bold=True)
        ws.cell(row=5, column=column_number + i - 1).alignment = Alignment(horizontal='center')

    for i, data in enumerate(transformed_data, start=6):
        ws.cell(row=i, column=column_number).value = data['identifier']
        ws.cell(row=i, column=column_number + 1).value = f"{data['strikePrice']} {'CE' if 'CE' in data['identifier'] else 'PE'}"
        ws.cell(row=i, column=column_number + 2).value = data['lastPrice']

    ws.cell(row=20, column=column_number).value = "NIFTY"
    ws.cell(row=20, column=column_number + 2).value = underlying_value
    print("[Created new .xlsx file]")

def add_new_lastprice(ws, transformed_data, timestamp, underlying_value, column_number):
    """Write transformed data to the worksheet."""
    last_col_index, last_col_value = find_last_column(ws)

    if last_col_value == timestamp.strftime('%H:%M'):
        print("[XLSX file already up-to-date]")
        return

    ws.cell(row=5, column=column_number + last_col_index - 1).value = timestamp.strftime('%H:%M')
    ws.cell(row=5, column=column_number + last_col_index - 1).font = Font(bold=True)
    ws.cell(row=5, column=column_number + last_col_index - 1).alignment = Alignment(horizontal='center')

    # Write data rows starting from row 6
    for i, data in enumerate(transformed_data, start=6):
        ws.cell(row=i, column=column_number + last_col_index - 1).value = data['lastPrice']

    ws.cell(row=20, column=column_number + last_col_index - 1).value = underlying_value
    print("[Updated saved .xlsx file]")

def find_last_column(ws):
    max_col = ws.max_column
    last_col_index = None
    last_col_value = None

    for col in range(1, max_col + 1):
        cell_value = ws.cell(row=5, column=col).value
        if cell_value is not None:
            last_col_index = col
            last_col_value = cell_value

    return last_col_index, last_col_value

def adjust_column_widths(ws):
    """Adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        column = col[1].column_letter  # Get the column letter of the first cell in the column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass  # Ignore TypeError for NoneType values

        adjusted_width = (max_length + 2) * 1.2  # Adding some padding
        ws.column_dimensions[column].width = adjusted_width