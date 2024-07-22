import requests
import argparse
import schedule
import time
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from sqlite_dumper import dump_data_to_sqlite, initialize_database

# Constants
SYMBOL = 'NIFTY'
TARGET_STRIKE_PRICES = [24200, 24300, 24400, 24500, 24600, 24700, 24700]
XLSX_FILE = f"report/{SYMBOL}_options_data.xlsx"

# Headers for HTTP request
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}

def format_datetime(time_obj, format = "%d-%b-%Y %H:%M:%S"):
    return datetime.strftime(time_obj, format)

def send_http_request():
    url = f"https://www.nseindia.com/api/option-chain-indices?symbol={SYMBOL}"
    session = requests.Session()
    response = session.get(url, headers=HEADERS)
    return response.json()

def get_rectified_data(time_now):
    print("\n=== Fetching data at", format_datetime(time_now) , "===")
    while True:
        data = send_http_request()
        last_timestamp_str = data['records']['timestamp']
        last_timestamp = datetime.strptime(last_timestamp_str, '%d-%b-%Y %H:%M:%S')

        # Calculate the time difference
        time_difference = time_now - last_timestamp

        # Check if the data is within the last minute
        if time_difference <= timedelta(seconds=59):
            print("Data found of", format_datetime(last_timestamp))
            print("Time difference:", time_difference)
            return data
        else:
            # Adjust the sleep time based on how often you want to check
            print("Data found was old. Retrying in 10s ...")
            time.sleep(10)

def fetch_option_chain(symbol, target_strike_prices):
    """Fetch option chain data from NSE website."""
    time_now = datetime.now()
    data = get_rectified_data(time_now)

    # Dump data to SQLite
    dump_data_to_sqlite(data, symbol)

    # Extract relevant data
    all_data = data['filtered']['data']
    underlying_value = data['records']['underlyingValue']
    timestamp_str = data['records']['timestamp']
    timestamp = datetime.strptime(timestamp_str, '%d-%b-%Y %H:%M:%S')

    # Filter data based on target_strike_prices
    filtered_options = [option for option in all_data if option.get("strikePrice") in target_strike_prices]

    return filtered_options, timestamp, underlying_value

def transform_data(filtered_options):
    """Transform raw option chain data into required format."""
    transformed_data = []

    for entry in filtered_options:
        ce_entry = {
            'strikePrice': entry['strikePrice'],
            'identifier': entry['CE']['identifier'],
            'lastPrice': entry['CE']['lastPrice']
        }
        transformed_data.append(ce_entry)

        pe_entry = {
            'strikePrice': entry['strikePrice'],
            'identifier': entry['PE']['identifier'],
            'lastPrice': entry['PE']['lastPrice']
        }
        transformed_data.append(pe_entry)

    return transformed_data

def update_xlsx(symbol, transformed_data, timestamp, underlying_value):
    """Update or create Excel file with option chain data."""
    try:
        wb = load_workbook(XLSX_FILE)
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active
    ws.title = 'Options Data'

    # Find the column where new data should start
    column_number = find_column_to_update(ws, timestamp)

    if column_number is None:
        column_number = find_empty_column(ws)
        initialize_sheet(ws, symbol, transformed_data, timestamp, underlying_value, column_number)
    else:
        add_new_lastprice(ws, transformed_data, timestamp, underlying_value, column_number)

    # Adjust column widths based on content
    adjust_column_widths(ws)

    # Save workbook
    wb.save(XLSX_FILE)
    print(f"XLSX file '{XLSX_FILE}' successfully updated.")

def find_column_to_update(ws, timestamp):
    """Find the column index to update based on timestamp."""
    max_col = ws.max_column
    for col in range(2, max_col + 1):
        cell_value = ws.cell(row=4, column=col).value
        if cell_value == f'DATE {timestamp.strftime("%d-%m-%Y")}':
            return col
    return None

def find_empty_column(ws):
    """Find the first empty column in the worksheet."""
    max_col = ws.max_column
    for col in range(2, max_col + 2):
        if ws.cell(row=3, column=col).value is None:
            return col
    return max_col + 1

def initialize_sheet(ws, symbol, transformed_data, timestamp, underlying_value, column_number):
    """Initialize the worksheet with headers and timestamp."""
    print("[CREATING NEW FILE]")
    ws.cell(row=3, column=column_number).value = symbol
    ws.cell(row=3, column=column_number).font = Font(bold=True)
    ws.cell(row=4, column=column_number).value = f'DATE {timestamp.strftime("%d-%m-%Y")}'
    ws.cell(row=4, column=column_number).font = Font(bold=True)

    headers = ['Identifier', 'Strike Price / Type', timestamp.strftime('%H:%M')]
    for i, header in enumerate(headers, start=1):
        ws.cell(row=5, column=column_number + i - 1).value = header
        ws.cell(row=5, column=column_number + i - 1).font = Font(bold=True)
        ws.cell(row=5, column=column_number + i - 1).alignment = Alignment(horizontal='center')

    for i, data in enumerate(transformed_data, start=6):
        ws.cell(row=i, column=column_number).value = data['identifier']
        ws.cell(row=i, column=column_number + 1).value = f"{data['strikePrice']} {'CE' if 'CE' in data['identifier'] else 'PE'}"
        ws.cell(row=i, column=column_number + 2).value = data['lastPrice']

    ws.cell(row=20, column=column_number).value = "NIFTY"
    ws.cell(row=20, column=column_number + 2).value = underlying_value

def add_new_lastprice(ws, transformed_data, timestamp, underlying_value, column_number):
    """Write transformed data to the worksheet."""
    print("[Updating saved file]")
    last_col_index, last_col_value = find_last_column(ws)

    if last_col_value == timestamp.strftime('%H:%M'):
        return

    ws.cell(row=5, column=column_number + last_col_index - 1).value = timestamp.strftime('%H:%M')
    ws.cell(row=5, column=column_number + last_col_index - 1).font = Font(bold=True)
    ws.cell(row=5, column=column_number + last_col_index - 1).alignment = Alignment(horizontal='center')

    # Write data rows starting from row 6
    for i, data in enumerate(transformed_data, start=6):
        ws.cell(row=i, column=column_number + last_col_index - 1).value = data['lastPrice']

    ws.cell(row=20, column=column_number + last_col_index - 1).value = underlying_value

def find_last_column(ws):
    max_col = ws.max_column
    last_col_index = None
    last_col_value = None

    for col in range(1, max_col + 1):
        cell_value = ws.cell(row=5, column=col).value
        if cell_value is not None:
            last_col_index = col
            last_col_value = cell_value

    return last_col_index, last_col_value

def adjust_column_widths(ws):
    """Adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        column = col[1].column_letter  # Get the column letter of the first cell in the column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass  # Ignore TypeError for NoneType values

        adjusted_width = (max_length + 2) * 1.2  # Adding some padding
        ws.column_dimensions[column].width = adjusted_width


def job():
    filtered_options, timestamp, underlying_value = fetch_option_chain(SYMBOL, TARGET_STRIKE_PRICES)
    transformed_data = transform_data(filtered_options)
    update_xlsx(SYMBOL, transformed_data, timestamp, underlying_value)

if __name__ == '__main__':
    initialize_database()

    parser = argparse.ArgumentParser(description="Run scheduled jobs for fetching and processing stock options data.")
    parser.add_argument('--init', '-i', action='store_true', help="Initialize the database and perform the initial fetch.")
    args = parser.parse_args()
    
    # Initial fetch and write to Excel if --init or -i
    if args.init:
        initialize_database()
        job()

    # List of times to run the job
    times = [
        "09:15", "09:30", "10:00", "10:30", "11:00", "11:30", "12:00",
        "12:30", "13:00", "13:30", "14:00", "14:30", "15:00", "15:35",
    ]

    # Schedule the job at the specified times
    for time_str in times:
        schedule.every().day.at(time_str + ":59").do(job)

    # Keep the script running
    while True:
        schedule.run_pending()
        time.sleep(1)
